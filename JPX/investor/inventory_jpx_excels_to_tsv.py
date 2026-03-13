# JPX/investor/inventory_jpx_excels_to_tsv.py
"""
File: JPX/investor/inventory_jpx_excels_to_tsv.py

ソースコードの役割:
本モジュールは、JPX（日本取引所グループ）が公表する投資部門別売買状況のExcelファイルから、
「Tokyo & Nagoya」シートの右ブロック（2市場計）の週次データを1レコードとして抽出し、TSV形式で出力します。
.xls形式の場合は、`xlrd` パッケージを用いて直接読み込みを行います。

主目的:
- 右ブロックの週を 1レコードとして出力する
- 外国人の売り・買い・差引を確実に抽出する
- 必要に応じて主要部門・内訳も同時に出力する

対応ファイル:
- .xls
- .xlsx
- .xlsm
"""

from __future__ import annotations

import argparse
import csv
import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

# ロギングの設定
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {".xls", ".xlsx", ".xlsm"}
TARGET_SHEET_NAME = "Tokyo & Nagoya"

# 右ブロック（2市場計）の抽出に必要なセル・列定義
RIGHT_WEEK_CANDIDATE_CELLS = ("H11", "I11", "H10", "I10")
RIGHT_BLOCK_COL_SHARES = "I"
RIGHT_BLOCK_COL_RATIO = "J"
RIGHT_BLOCK_COL_BALANCE = "K"


@dataclass(frozen=True)
class SectionSpec:
    """抽出対象となる各投資部門（セクション）の定義情報を保持するデータクラス。

    Attributes:
        key_prefix: 抽出結果の辞書キーに使用するプレフィックス（例: "foreigners"）。
        labels: 行を特定するための検索ラベル群。
        anchor_kind: 基準となる行の種類（"purchases", "sales", "total"）。
        row_range: ラベル検索対象の行範囲。
    """
    key_prefix: str
    labels: Tuple[str, ...]
    anchor_kind: str
    row_range: Optional[Tuple[int, int]] = None


# 抽出対象の投資部門の仕様リスト
SECTION_SPECS: Tuple[SectionSpec, ...] = (
    SectionSpec("proprietary", ("Proprietary",), "purchases", (13, 15)),
    SectionSpec("institutions", ("Institutions",), "purchases", (24, 26)),
    SectionSpec("individuals", ("Individuals",), "purchases", (27, 29)),
    SectionSpec("foreigners", ("Foreigners",), "purchases", (30, 32)),
    SectionSpec("securities_cos", ("Securities Cos.",), "purchases", (33, 35)),
    SectionSpec("trusts", ("Trusts",), "total", (38, 40)),
    SectionSpec("business_corps", ("Business Cos.",), "purchases", (41, 43)),
    SectionSpec("other_corps", ("Other Cos.",), "purchases", (44, 46)),
    SectionSpec("financial_institutions", ("Financial Institutions", "Institutions"), "total", (47, 49)),
    SectionSpec("life_and_non_life", ("Life & Non-Life",), "purchases", (52, 54)),
    SectionSpec("city_and_regional_banks", ("City & Regional BK",), "purchases", (55, 57)),
    SectionSpec("trust_banks", ("Trust BK",), "purchases", (58, 60)),
    SectionSpec("other_financials", ("Other Financials",), "purchases", (61, 63)),
)


def parse_args() -> argparse.Namespace:
    """コマンドライン引数をパースします。

    Returns:
        argparse.Namespace: パース済みの引数オブジェクト。
    """
    parser = argparse.ArgumentParser(
        description="Tokyo & Nagoya シート右ブロックを 1週 1レコードで TSV 化する"
    )
    parser.add_argument(
        "--input-dir", 
        default="C:\\transformer_futures_data\\excel\\InvestorType",
        help="入力ディレクトリ"
    )
    parser.add_argument(
        "--output-tsv", 
        default="C:\\transformer_futures_data\\tsv\\InvestorType\\jpx_investor_type_inventory.tsv",
        help="出力TSVファイル"
    )
    parser.add_argument(
        "--sheet",
        default=TARGET_SHEET_NAME,
        help=f"対象シート名 (default: {TARGET_SHEET_NAME})",
    )
    parser.add_argument(
        "--recursive",
        default=True,
        action="store_true",
        help="入力ディレクトリ配下を再帰的に探索する",
    )
    return parser.parse_args()


def iter_excel_files(input_dir: Path, recursive: bool) -> Iterable[Path]:
    """指定されたディレクトリ内のExcelファイルを探索するジェネレータ。

    Args:
        input_dir: 探索起点となるディレクトリ。
        recursive: 再帰的に探索するかどうか。

    Yields:
        Path: 見つかったサポート対象のExcelファイルのパス。
    """
    pattern = "**/*" if recursive else "*"
    for path in sorted(input_dir.glob(pattern)):
        if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS:
            yield path


class XlrdWorksheetWrapper:
    """xlrdのシートをopenpyxlのWorksheetインターフェースに適合させるラッパー。"""
    def __init__(self, sheet: Any):
        self.sheet = sheet
        self.max_row = sheet.nrows

    def __getitem__(self, key: str) -> Any:
        """セルの座標（例: 'A1'）からセルの値をラップして返します。"""
        col_str, row = coordinate_from_string(key)
        col = column_index_from_string(col_str)
        
        # openpyxlは1-based、xlrdは0-basedなので変換
        row_idx = row - 1
        col_idx = col - 1

        class DummyCell:
            def __init__(self, value: Any):
                self.value = value

        # シートの範囲外を参照した場合はNoneを持つセルを返す
        if row_idx >= self.sheet.nrows or col_idx >= self.sheet.ncols:
            return DummyCell(None)
            
        val = self.sheet.cell_value(row_idx, col_idx)
        # xlrdは空セルを空文字列で返すことがあるため、Noneに正規化
        if val == "":
            val = None
        return DummyCell(val)


class XlrdWorkbookWrapper:
    """xlrdのブックをopenpyxlのWorkbookインターフェースに適合させるラッパー。"""
    def __init__(self, path: Path):
        try:
            import xlrd
        except ImportError:
            raise ImportError(
                "'.xls' ファイルを読み込むには 'xlrd' パッケージが必要です。"
                "'pip install xlrd' を実行してください。"
            )
        self.wb = xlrd.open_workbook(str(path))
        self.sheetnames = self.wb.sheet_names()

    def __getitem__(self, sheet_name: str) -> XlrdWorksheetWrapper:
        """シート名でシートオブジェクトを取得します。"""
        return XlrdWorksheetWrapper(self.wb.sheet_by_name(sheet_name))

    def close(self) -> None:
        """リソースを解放します（xlrdでは実質何もしない）。"""
        pass


def load_workbook_for_any_excel(path: Path) -> Any:
    """拡張子に応じてExcelファイルを適切に読み込みます。

    .xlsx は openpyxl で、.xls は xlrd で読み込み、共通のインターフェースで返します。

    Args:
        path: 読み込み対象のExcelファイルのパス。

    Returns:
        Any: 読み込まれたワークブック（openpyxl.Workbook または XlrdWorkbookWrapper）。

    Raises:
        ValueError: 未対応の拡張子の場合。
    """
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return load_workbook(path, data_only=True, read_only=True)

    if suffix == ".xls":
        return XlrdWorkbookWrapper(path)

    raise ValueError(f"未対応の拡張子です: {path}")


def normalize_text(value: Any) -> str:
    """セルの値を空白文字を取り除いた文字列として正規化します。
    
    Args:
        value (Any): Excelセルから取得した値。
        
    Returns:
        str: 正規化された文字列。Noneの場合は空文字列。
    """
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def normalize_number_text(value: Any) -> str:
    """数値表現からカンマやパーセント記号を除去して正規化します。
    
    Args:
        value (Any): Excelセルから取得した数値や文字列。
        
    Returns:
        str: 数値変換用にクレンジングされた文字列。
    """
    text = normalize_text(value)
    if not text:
        return ""
    text = text.replace(",", "").replace("%", "")
    return text


def get_cell_text(ws: Any, cell_ref: str) -> str:
    """指定されたセル参照から正規化されたテキストを取得します。
    
    Args:
        ws (Any): 対象のワークシート。
        cell_ref (str): セルのアドレス（例: "A1"）。
        
    Returns:
        str: 取得・正規化されたテキスト。
    """
    return normalize_text(ws[cell_ref].value)


def find_label_row(
    ws: Any,
    labels: Sequence[str],
    row_range: Optional[Tuple[int, int]] = None,
) -> Optional[int]:
    """A列内で指定されたラベルを含む行番号を探索します。

    Args:
        ws: 対象のワークシート。
        labels: 探索するラベル文字列のリスト。
        row_range: 探索する行の範囲 (開始行, 終了行)。

    Returns:
        Optional[int]: 見つかった場合は行番号。見つからない場合はNone。
    """
    start_row = row_range[0] if row_range else 1
    end_row = row_range[1] if row_range else ws.max_row
    normalized_labels = tuple(label.strip().lower() for label in labels)

    for row_idx in range(start_row, end_row + 1):
        a_value = normalize_text(ws[f"A{row_idx}"].value).lower()
        if not a_value:
            continue
        if any(a_value == label or label in a_value for label in normalized_labels):
            return row_idx
    return None


def resolve_triplet_rows(anchor_row: int, anchor_kind: str) -> Tuple[int, int, int]:
    """基準となる行（売り・買い・合計）から、(売り行, 買い行, 合計行) の3行を特定します。
    
    Args:
        anchor_row (int): 見つかった基準ラベルの行番号。
        anchor_kind (str): ラベルが示す行の種類（"sales", "purchases", "total"）。
        
    Returns:
        Tuple[int, int, int]: (売り行, 買い行, 合計行) の行番号。
        
    Raises:
        ValueError: 未知の anchor_kind が指定された場合。
    """
    if anchor_kind == "sales":
        return anchor_row, anchor_row + 1, anchor_row + 2
    if anchor_kind == "purchases":
        return anchor_row - 1, anchor_row, anchor_row + 1
    if anchor_kind == "total":
        return anchor_row - 2, anchor_row - 1, anchor_row
    raise ValueError(f"未知の anchor_kind です: {anchor_kind}")


def first_non_empty(*values: str) -> str:
    """与えられた引数の中から、最初の空ではない文字列を返します。
    
    Args:
        *values (str): 評価する文字列の可変長引数。
        
    Returns:
        str: 最初に見つかった空でない文字列。すべて空の場合は空文字列。
    """
    for value in values:
        if value:
            return value
    return ""


def extract_section(ws: Any, spec: SectionSpec) -> Dict[str, str]:
    """1つのセクション（投資部門等）の売り・買い・合計・差引に関するデータを抽出します。

    Args:
        ws: 対象のワークシート。
        spec: 抽出するセクションの仕様情報。

    Returns:
        Dict[str, str]: セクション情報をパースした辞書。
    """
    row_idx = find_label_row(ws, spec.labels, spec.row_range)
    if row_idx is None:
        logger.warning("ラベル行が見つかりませんでした: %s", ", ".join(spec.labels))
        return {
            f"{spec.key_prefix}_sales": "",
            f"{spec.key_prefix}_sales_ratio": "",
            f"{spec.key_prefix}_purchases": "",
            f"{spec.key_prefix}_purchases_ratio": "",
            f"{spec.key_prefix}_total": "",
            f"{spec.key_prefix}_total_ratio": "",
            f"{spec.key_prefix}_balance": "",
        }

    sales_row, purchases_row, total_row = resolve_triplet_rows(row_idx, spec.anchor_kind)

    # 共通のカラム位置を使用してデータを取得
    sales_value = normalize_number_text(ws[f"{RIGHT_BLOCK_COL_SHARES}{sales_row}"].value)
    sales_ratio = normalize_number_text(ws[f"{RIGHT_BLOCK_COL_RATIO}{sales_row}"].value)
    sales_balance = normalize_number_text(ws[f"{RIGHT_BLOCK_COL_BALANCE}{sales_row}"].value)

    purchases_value = normalize_number_text(ws[f"{RIGHT_BLOCK_COL_SHARES}{purchases_row}"].value)
    purchases_ratio = normalize_number_text(ws[f"{RIGHT_BLOCK_COL_RATIO}{purchases_row}"].value)
    purchases_balance = normalize_number_text(ws[f"{RIGHT_BLOCK_COL_BALANCE}{purchases_row}"].value)

    total_value = normalize_number_text(ws[f"{RIGHT_BLOCK_COL_SHARES}{total_row}"].value)
    total_ratio = normalize_number_text(ws[f"{RIGHT_BLOCK_COL_RATIO}{total_row}"].value)
    total_balance = normalize_number_text(ws[f"{RIGHT_BLOCK_COL_BALANCE}{total_row}"].value)

    # 差引は売り・買い・合計のいずれかの行に記録されているため、存在するものを採用
    balance_value = first_non_empty(purchases_balance, sales_balance, total_balance)

    return {
        f"{spec.key_prefix}_sales": sales_value,
        f"{spec.key_prefix}_sales_ratio": sales_ratio,
        f"{spec.key_prefix}_purchases": purchases_value,
        f"{spec.key_prefix}_purchases_ratio": purchases_ratio,
        f"{spec.key_prefix}_total": total_value,
        f"{spec.key_prefix}_total_ratio": total_ratio,
        f"{spec.key_prefix}_balance": balance_value,
    }


WEEK_PATTERN = re.compile(r"\d{2}/\d{2}.*\d{2}/\d{2}")


def extract_right_week_period(ws: Any) -> str:
    """右ブロックから対象期間（週）を示す文字列を抽出します。
    
    Args:
        ws (Any): 対象のワークシート。
        
    Returns:
        str: 抽出された週期間の文字列（例: "01/05 ~ 01/09"）。見つからない場合は空文字列。
    """
    for cell_ref in RIGHT_WEEK_CANDIDATE_CELLS:
        text = get_cell_text(ws, cell_ref)
        if text:
            return text

    # 指定セルになければヘッダ行付近をフォールバックとして検索
    for row_idx in range(1, min(ws.max_row, 20) + 1):
        for col in ("H", "I", "J", "K"):
            text = normalize_text(ws[f"{col}{row_idx}"].value)
            if text and WEEK_PATTERN.search(text):
                return text
    return ""


def extract_record(path: Path, input_dir: Path, sheet_name: str) -> Optional[Dict[str, str]]:
    """1つのExcelファイルからTSVに出力する1レコード分の情報を抽出します。

    Args:
        path: Excelファイルのパス。
        input_dir: ベースディレクトリ。
        sheet_name: 対象シート名。

    Returns:
        Optional[Dict[str, str]]: 抽出されたレコード。失敗時はNone。
    """
    try:
        workbook = load_workbook_for_any_excel(path)
    except Exception as exc:
        logger.error("Excelファイルの読み込みに失敗しました (%s): %s", path.name, exc)
        return None

    try:
        if sheet_name not in workbook.sheetnames:
            logger.info("対象シートが存在しないためスキップします: %s (%s)", path.name, sheet_name)
            return None

        ws = workbook[sheet_name]
        record: Dict[str, str] = {
            "file_name": path.name,
            "relative_path": str(path.relative_to(input_dir)),
            "week_period": extract_right_week_period(ws),
        }

        # 定義された全セクションについてデータをパースしてレコードに追加
        for spec in SECTION_SPECS:
            record.update(extract_section(ws, spec))

        return record
    finally:
        workbook.close()


# TSV出力用のカラム定義（外国人データを先頭に配置）
FIELDNAMES = [
    "file_name",                                # ファイル名
    "relative_path",                            # 入力ディレクトリからの相対パス
    "week_period",                              # 対象週（例: 03/02~03/06）
    "foreigners_sales",                         # 外国人_売り_株数(または金額)
    "foreigners_sales_ratio",                   # 外国人_売り_比率
    "foreigners_purchases",                     # 外国人_買い_株数(または金額)
    "foreigners_purchases_ratio",               # 外国人_買い_比率
    "foreigners_total",                         # 外国人_合計_株数(または金額)
    "foreigners_total_ratio",                   # 外国人_合計_比率
    "foreigners_balance",                       # 外国人_差引
    "proprietary_sales",                        # 自己_売り_株数(または金額)
    "proprietary_sales_ratio",                  # 自己_売り_比率
    "proprietary_purchases",                    # 自己_買い_株数(または金額)
    "proprietary_purchases_ratio",              # 自己_買い_比率
    "proprietary_total",                        # 自己_合計_株数(または金額)
    "proprietary_total_ratio",                  # 自己_合計_比率
    "proprietary_balance",                      # 自己_差引
    "institutions_sales",                       # 法人_売り_株数(または金額)
    "institutions_sales_ratio",                 # 法人_売り_比率
    "institutions_purchases",                   # 法人_買い_株数(または金額)
    "institutions_purchases_ratio",             # 法人_買い_比率
    "institutions_total",                       # 法人_合計_株数(または金額)
    "institutions_total_ratio",                 # 法人_合計_比率
    "institutions_balance",                     # 法人_差引
    "individuals_sales",                        # 個人_売り_株数(または金額)
    "individuals_sales_ratio",                  # 個人_売り_比率
    "individuals_purchases",                    # 個人_買い_株数(または金額)
    "individuals_purchases_ratio",              # 個人_買い_比率
    "individuals_total",                        # 個人_合計_株数(または金額)
    "individuals_total_ratio",                  # 個人_合計_比率
    "individuals_balance",                      # 個人_差引
    "securities_cos_sales",                     # 証券会社_売り_株数(または金額)
    "securities_cos_sales_ratio",               # 証券会社_売り_比率
    "securities_cos_purchases",                 # 証券会社_買い_株数(または金額)
    "securities_cos_purchases_ratio",           # 証券会社_買い_比率
    "securities_cos_total",                     # 証券会社_合計_株数(または金額)
    "securities_cos_total_ratio",               # 証券会社_合計_比率
    "securities_cos_balance",                   # 証券会社_差引
    "trusts_sales",                             # 投資信託_売り_株数(または金額)
    "trusts_sales_ratio",                       # 投資信託_売り_比率
    "trusts_purchases",                         # 投資信託_買い_株数(または金額)
    "trusts_purchases_ratio",                   # 投資信託_買い_比率
    "trusts_total",                             # 投資信託_合計_株数(または金額)
    "trusts_total_ratio",                       # 投資信託_合計_比率
    "trusts_balance",                           # 投資信託_差引
    "business_corps_sales",                     # 事業法人_売り_株数(または金額)
    "business_corps_sales_ratio",               # 事業法人_売り_比率
    "business_corps_purchases",                 # 事業法人_買い_株数(または金額)
    "business_corps_purchases_ratio",           # 事業法人_買い_比率
    "business_corps_total",                     # 事業法人_合計_株数(または金額)
    "business_corps_total_ratio",               # 事業法人_合計_比率
    "business_corps_balance",                   # 事業法人_差引
    "other_corps_sales",                        # その他法人等_売り_株数(または金額)
    "other_corps_sales_ratio",                  # その他法人等_売り_比率
    "other_corps_purchases",                    # その他法人等_買い_株数(または金額)
    "other_corps_purchases_ratio",              # その他法人等_買い_比率
    "other_corps_total",                        # その他法人等_合計_株数(または金額)
    "other_corps_total_ratio",                  # その他法人等_合計_比率
    "other_corps_balance",                      # その他法人等_差引
    "financial_institutions_sales",             # 金融機関_売り_株数(または金額)
    "financial_institutions_sales_ratio",       # 金融機関_売り_比率
    "financial_institutions_purchases",         # 金融機関_買い_株数(または金額)
    "financial_institutions_purchases_ratio",   # 金融機関_買い_比率
    "financial_institutions_total",             # 金融機関_合計_株数(または金額)
    "financial_institutions_total_ratio",       # 金融機関_合計_比率
    "financial_institutions_balance",           # 金融機関_差引
    "life_and_non_life_sales",                  # 生保・損保_売り_株数(または金額)
    "life_and_non_life_sales_ratio",            # 生保・損保_売り_比率
    "life_and_non_life_purchases",              # 生保・損保_買い_株数(または金額)
    "life_and_non_life_purchases_ratio",        # 生保・損保_買い_比率
    "life_and_non_life_total",                  # 生保・損保_合計_株数(または金額)
    "life_and_non_life_total_ratio",            # 生保・損保_合計_比率
    "life_and_non_life_balance",                # 生保・損保_差引
    "city_and_regional_banks_sales",            # 都銀・地銀等_売り_株数(または金額)
    "city_and_regional_banks_sales_ratio",      # 都銀・地銀等_売り_比率
    "city_and_regional_banks_purchases",        # 都銀・地銀等_買い_株数(または金額)
    "city_and_regional_banks_purchases_ratio",  # 都銀・地銀等_買い_比率
    "city_and_regional_banks_total",            # 都銀・地銀等_合計_株数(または金額)
    "city_and_regional_banks_total_ratio",      # 都銀・地銀等_合計_比率
    "city_and_regional_banks_balance",          # 都銀・地銀等_差引
    "trust_banks_sales",                        # 信託銀行_売り_株数(または金額)
    "trust_banks_sales_ratio",                  # 信託銀行_売り_比率
    "trust_banks_purchases",                    # 信託銀行_買い_株数(または金額)
    "trust_banks_purchases_ratio",              # 信託銀行_買い_比率
    "trust_banks_total",                        # 信託銀行_合計_株数(または金額)
    "trust_banks_total_ratio",                  # 信託銀行_合計_比率
    "trust_banks_balance",                      # 信託銀行_差引
    "other_financials_sales",                   # その他金融機関_売り_株数(または金額)
    "other_financials_sales_ratio",             # その他金融機関_売り_比率
    "other_financials_purchases",               # その他金融機関_買い_株数(または金額)
    "other_financials_purchases_ratio",         # その他金融機関_買い_比率
    "other_financials_total",                   # その他金融機関_合計_株数(または金額)
    "other_financials_total_ratio",             # その他金融機関_合計_比率
    "other_financials_balance",                 # その他金融機関_差引
]


def write_tsv(records: Sequence[Dict[str, str]], output_path: Path) -> None:
    """抽出されたレコード群をTSVファイルに出力します。

    Args:
        records: 書き出す対象のレコード（辞書形式）のリスト。
        output_path: 出力先となるTSVファイルのパス。
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES, delimiter="\t", extrasaction="ignore")
        writer.writeheader()
        for record in records:
            writer.writerow(record)
    logger.info("TSVへの書き込みが完了しました: %s", output_path)


def main() -> None:
    """スクリプトのメインエントリポイント。
    
    引数をパースし、指定されたディレクトリ内のExcelファイルを探索・変換・抽出し、
    最終的な結果をTSVファイルとして出力します。
    """
    args = parse_args()
    input_dir = Path(args.input_dir)
    output_path = Path(args.output_tsv)

    if not input_dir.exists() or not input_dir.is_dir():
        logger.error("入力ディレクトリが存在しません: %s", input_dir)
        return

    records = []
    for excel_path in iter_excel_files(input_dir, recursive=args.recursive):
        logger.info("処理中: %s", excel_path.name)
        record = extract_record(excel_path, input_dir, args.sheet)
        if record is not None:
            records.append(record)

    if records:
        write_tsv(records, output_path)
    else:
        logger.warning("出力対象のデータが見つかりませんでした。")


if __name__ == "__main__":
    main()
